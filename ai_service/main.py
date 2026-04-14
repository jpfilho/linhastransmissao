import os
import io
import json
from typing import Optional, List
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from supabase import create_client, Client
import openpyxl
import logging

logging.basicConfig(
    filename='ai_service.log',
    level=logging.ERROR,
    format='%(asctime)s %(levelname)s: %(message)s'
)

from models import (
    AnalyzeImageRequest,
    AnalyzeImageResponse,
    CompareImagesRequest,
    CompareImagesResponse,
    GenerateSummaryRequest,
    GenerateSummaryResponse,
    HealthResponse,
    ParseRocoRequest,
    ParseRocoResponse,
    RocoSegmentModel,
)
from analyzer import ImageAnalyzer

load_dotenv()

# Configuration
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
SUPABASE_URL = os.getenv("SUPABASE_URL", "http://127.0.0.1:54321")
SUPABASE_SERVICE_KEY = os.getenv("SUPABASE_SERVICE_KEY", "")
MOONDREAM_API_KEY = os.getenv("MOONDREAM_API_KEY", "")

# Global instances
analyzer: Optional[ImageAnalyzer] = None
supabase: Optional[Client] = None


@asynccontextmanager
async def lifespan(app: FastAPI):
    global analyzer, supabase
    analyzer = ImageAnalyzer(openai_api_key=OPENAI_API_KEY, moondream_api_key=MOONDREAM_API_KEY)
    if SUPABASE_SERVICE_KEY:
        supabase = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
    ok = lambda v: "[OK]" if v else "[--]"
    print(f"AI Service started | OpenAI: {ok(OPENAI_API_KEY)} | Supabase: {ok(supabase)} | Moondream: {ok(MOONDREAM_API_KEY)}")

    yield
    print("AI Service shutting down")


app = FastAPI(
    title="Inspeção Aérea - AI Service",
    description="AI analysis for transmission line aerial inspection photos",
    version="1.0.0",
    lifespan=lifespan,
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["X-Tower-Code", "X-Tower-Function", "X-Tower-Structure", "X-Tower-Height", "X-AI-Detection"],
)


@app.get("/health", response_model=HealthResponse)
async def health_check():
    return HealthResponse(
        status="ok",
        version="1.0.0",
        services={
            "openai": bool(OPENAI_API_KEY),
            "supabase": supabase is not None,
        },
    )


@app.post("/analyze-image", response_model=AnalyzeImageResponse)
async def analyze_image(request: AnalyzeImageRequest):
    """Analyze a single aerial inspection photo using CV + GPT-4o."""
    if not analyzer:
        raise HTTPException(status_code=503, detail="Analyzer not initialized")

    try:
        result = await analyzer.analyze_image(request.image_url, request.photo_id)

        # Save to Supabase
        if supabase:
            try:
                supabase.table("ai_analysis").upsert({
                    "foto_id": request.photo_id,
                    "vegetation_detected": result.vegetation_detected,
                    "vegetation_score": result.vegetation_score,
                    "fire_signs": result.fire_signs,
                    "fire_score": result.fire_score,
                    "structural_issue": result.structural_issue,
                    "anomaly_type": result.anomaly_type,
                    "severity_score": result.severity_score,
                    "confidence": result.confidence,
                    "quality_blur": result.quality.blur_score,
                    "quality_exposure": result.quality.exposure_score,
                    "bounding_boxes": [bb.model_dump() for bb in result.bounding_boxes],
                    "summary": result.summary,
                    "model_version": result.model_version,
                }, on_conflict="foto_id").execute()
            except Exception as e:
                print(f"Failed to save analysis to DB: {e}")

        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/annotate-image")
async def annotate_image(request: AnalyzeImageRequest):
    """Analyze image, identify tower type/function, and return annotated image with tower name written on it."""
    if not analyzer:
        raise HTTPException(status_code=503, detail="Analyzer not initialized")

    try:
        # Get tower code and altitude from DB
        torre_codigo = ""
        altitude = None
        if supabase:
            try:
                foto = supabase.table("fotos").select(
                    "torre_id, altitude, torres(codigo_torre)"
                ).eq("id", request.photo_id).maybe_single().execute()
                if foto.data:
                    if foto.data.get("torres") and foto.data["torres"].get("codigo_torre"):
                        torre_codigo = foto.data["torres"]["codigo_torre"]
                    altitude = foto.data.get("altitude")
            except Exception as e:
                print(f"[annotate-image] DB query error: {e}")

        # Download and analyze image
        img = await analyzer.download_image(request.image_url)
        
        # Build context with altitude info for height estimation
        context = torre_codigo
        if altitude:
            context = f"{torre_codigo} (foto tirada a {altitude}m de altitude)"
        
        gpt_result = await analyzer.analyze_with_gpt4o(img, request.photo_id, torre_codigo=context)

        # Extract tower identification
        tower_id = gpt_result.get("tower_identification", {})
        tower_function = tower_id.get("tower_function", "desconhecido")
        tower_structure = tower_id.get("tower_structure", "desconhecido")
        height_m = tower_id.get("height_estimate_m")
        
        # Build annotation label
        function_labels = {
            "ancoragem": "Ancoragem",
            "suspensao": "Suspensão",
            "transposicao": "Transposição",
            "derivacao": "Derivação",
            "terminal": "Terminal",
        }
        function_label = function_labels.get(tower_function, tower_function)
        
        # Get severity from scene summary
        severity = "none"
        scene = gpt_result.get("scene_summary", {})
        if scene.get("fire_present"):
            severity = "critical"
        elif scene.get("vegetation_present"):
            severity = "medium"

        # Generate annotated image with function + structure
        label = torre_codigo or tower_id.get("tower_code_visible") or "Desconhecida"
        type_label = f"{function_label} - {tower_structure.replace('_', ' ').title()}"
        if height_m:
            type_label += f" (~{height_m}m)"
        annotated_bytes = analyzer.annotate_image(img, label, tower_type=type_label, severity=severity)

        from fastapi.responses import Response
        return Response(
            content=annotated_bytes,
            media_type="image/jpeg",
            headers={
                "X-Tower-Code": label,
                "X-Tower-Function": tower_function,
                "X-Tower-Structure": tower_structure,
                "X-Tower-Height": str(height_m) if height_m else "0",
                "X-AI-Detection": json.dumps(gpt_result, ensure_ascii=False)[:2000],
            },
        )
    except Exception as e:
        print(f"[annotate-image] ERROR: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/compare-images", response_model=CompareImagesResponse)
async def compare_images(request: CompareImagesRequest):
    """Compare two images for temporal change detection."""
    if not analyzer:
        raise HTTPException(status_code=503, detail="Analyzer not initialized")

    try:
        result = await analyzer.compare_images(
            request.current_image_url,
            request.previous_image_url,
            request.photo_atual_id,
            request.photo_anterior_id,
        )

        # Save to Supabase
        if supabase:
            try:
                supabase.table("ai_comparisons").insert({
                    "foto_atual_id": request.photo_atual_id,
                    "foto_anterior_id": request.photo_anterior_id,
                    "torre_id": request.torre_id,
                    "change_detected": result.change_detected,
                    "vegetation_growth_level": result.vegetation_growth_level,
                    "degradation_level": result.degradation_level,
                    "new_anomaly_detected": result.new_anomaly_detected,
                    "comparison_details": result.comparison_details,
                }).execute()
            except Exception as e:
                print(f"Failed to save comparison to DB: {e}")

        return result
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/generate-summary", response_model=GenerateSummaryResponse)
async def generate_summary(request: GenerateSummaryRequest):
    """Generate LLM-powered inspection summary."""
    if not analyzer:
        raise HTTPException(status_code=503, detail="Analyzer not initialized")

    try:
        # If we have analysis data from the DB, pass it along
        analysis_data = request.analysis_data
        if not analysis_data and supabase:
            try:
                existing = supabase.table("ai_analysis").select("*").eq("foto_id", request.photo_id).maybe_single().execute()
                if existing and existing.data:
                    analysis_data = existing.data
            except Exception:
                pass

        result = await analyzer.generate_summary(
            request.photo_id,
            request.image_url,
            analysis_data,
        )

        # Only save to DB if it's a real report (not an error)
        if supabase and not result.content.startswith("Erro"):
            try:
                # Delete old reports for this photo first
                supabase.table("ai_reports").delete().eq("foto_id", request.photo_id).execute()
                supabase.table("ai_reports").insert({
                    "foto_id": request.photo_id,
                    "report_type": "photo_analysis",
                    "content": result.content,
                    "suggested_action": result.suggested_action,
                    "risk_interpretation": result.risk_interpretation,
                    "model_used": result.model_used,
                }).execute()
            except Exception as e:
                print(f"Failed to save report to DB: {e}")

        return result
    except Exception as e:
        print(f"Generate summary error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


from pydantic import BaseModel

class AnalyzeSupressaoRequest(BaseModel):
    torre_id: str
    photo_id: Optional[str] = None  # specific photo, or latest if None
    supressao_id: Optional[str] = None  # specific mapping record, or first for tower


@app.post("/analyze-supressao")
async def analyze_supressao(request: AnalyzeSupressaoRequest):
    """Analyze photo vs vegetation suppression mapping for a tower."""
    if not analyzer or not supabase:
        raise HTTPException(status_code=503, detail="Service not ready")

    try:
        # Get suppression data for this tower
        if request.supressao_id:
            supp_resp = supabase.table("mapeamento_supressao").select("*").eq("id", request.supressao_id).execute()
        else:
            supp_resp = supabase.table("mapeamento_supressao").select("*").eq("torre_id", request.torre_id).execute()
        
        if not supp_resp.data:
            raise HTTPException(status_code=404, detail="Nenhum dado de supressão encontrado para esta torre")
        
        suppression_data = supp_resp.data[0]  # Use first/specified record
        
        # Debug: log all non-null suppression data fields
        print(f"[ANALYZE-SUPRESSAO] tower={request.torre_id}")
        key_fields = ['est_codigo', 'vao_frente_m', 'largura_m', 'map_mec_extensao', 'map_mec_largura',
                       'map_man_extensao', 'map_man_largura', 'exec_mec_extensao', 'exec_man_extensao',
                       'roco_concluido', 'prioridade', 'descricao_servico', 'atende']
        for f in key_fields:
            v = suppression_data.get(f)
            print(f"  {f}: {v}")

        # Get tower info
        torre_resp = supabase.table("torres").select("id, codigo_torre, linha_id").eq("id", request.torre_id).execute()
        torre_codigo = torre_resp.data[0]["codigo_torre"] if torre_resp.data else ""

        # Get photo for this tower
        if request.photo_id:
            photo_resp = supabase.table("fotos").select("id, caminho_storage, torre_id").eq("id", request.photo_id).execute()
        else:
            photo_resp = supabase.table("fotos").select("id, caminho_storage, torre_id").eq("torre_id", request.torre_id).order("criado_em", desc=True).limit(1).execute()
        
        if not photo_resp.data:
            raise HTTPException(status_code=404, detail="Nenhuma foto encontrada para esta torre")
        
        photo = photo_resp.data[0]
        # Build public URL from Supabase Storage
        storage_path = photo["caminho_storage"]
        image_url = f"{SUPABASE_URL}/storage/v1/object/public/fotos-inspecao/{storage_path}"
        print(f"[ANALYZE-SUPRESSAO] photo={photo['id']}, url={image_url[:80]}...")

        # Run AI analysis
        result = await analyzer.analyze_with_suppression(
            image_url=image_url,
            photo_id=photo["id"],
            suppression_data=suppression_data,
            torre_codigo=torre_codigo,
        )

        return {"success": True, **result}

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/import-supressao")
async def import_supressao(
    file: UploadFile = File(...),
    campanha_roco: str = Form("2026"),
    linha_id: Optional[str] = Form(None),
):
    """Import vegetation suppression mapping from Excel worksheet."""
    if not supabase:
        raise HTTPException(status_code=503, detail="Supabase not connected")
    
    # Handle empty string from form
    if linha_id is not None and linha_id.strip() == "":
        linha_id = None

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active

        # Extract LT name from cell D2 or use provided linha_id
        lt_name_cell = ws.cell(row=2, column=4).value or ""
        # e.g. "LT: SIJ / BEA 05C6" -> extract useful part
        lt_name = str(lt_name_cell).replace("LT:", "").strip()

        # If linha_id not provided, try to find by name
        if not linha_id:
            lines = supabase.table("linhas").select("id, nome, codigo").execute()
            
            # Normalize function: "SIJ / BEA 05C6" -> "sjibea"
            import re
            def normalize_lt(name):
                # Remove common prefixes/suffixes and codes like 05C6, O5C9, W1, W2
                n = name.upper()
                n = re.sub(r'\b(LT|CE|W\d+|[O0]\d[A-Z]\d|MAPEAMENTO|\d{4}|LINK)\b', '', n)
                n = re.sub(r'[^A-Z]', '', n)  # keep only letters
                return n.lower()
            
            lt_norm = normalize_lt(lt_name)
            # Also try sheet name as backup
            sheet_norm = normalize_lt(ws.title or "")
            
            for l in lines.data:
                line_norm = normalize_lt(l["nome"])
                code_norm = normalize_lt(l.get("codigo") or "")
                # Check if the normalized names overlap enough
                if (lt_norm and (lt_norm in line_norm or line_norm in lt_norm)) or \
                   (sheet_norm and (sheet_norm in line_norm or line_norm in sheet_norm)) or \
                   (lt_norm and (lt_norm in code_norm or code_norm in lt_norm)):
                    linha_id = l["id"]
                    break

        # Build tower code lookup for matching EST codes
        tower_map = {}  # EST code -> torre_id
        if linha_id:
            towers = supabase.table("torres").select("id, codigo_torre").eq("linha_id", linha_id).limit(5000).execute()
            for t in towers.data:
                code = t["codigo_torre"]
                # Normalize: "SJIBEAW1 2-1" -> "2/1", "TGDSBTW1 018-1" -> "18/1"
                parts = code.split(" ")
                if len(parts) >= 2:
                    est = parts[-1].replace("-", "/")
                    tower_map[est] = t["id"]
                    # Also strip leading zeros: "018/1" -> "18/1"
                    est_parts = est.split("/")
                    if len(est_parts) == 2:
                        stripped = est_parts[0].lstrip("0") + "/" + est_parts[1]
                        tower_map[stripped] = t["id"]
                # Also try the code as-is with - replaced
                tower_map[code.replace("-", "/")] = t["id"]
            print(f"[IMPORT] linha_id={linha_id}, towers found={len(towers.data)}, tower_map size={len(tower_map)}")
            if tower_map:
                print(f"[IMPORT] Sample tower map keys: {list(tower_map.keys())[:10]}")
        else:
            print(f"[IMPORT] WARNING: linha_id is None, no tower matching will occur")

        # Parse data rows (starting from row 6)
        results = []
        imported = 0
        skipped = 0
        errors = 0

        for row_idx in range(6, ws.max_row + 1):
            est = ws.cell(row=row_idx, column=1).value
            if est is None:
                continue
            est = str(est).strip()
            if not est or est in ("TOTAL", "Subtotal", "PORT.", "TGD"):
                continue

            # Get cell values
            def num(col):
                v = ws.cell(row=row_idx, column=col).value
                if v is None:
                    return None
                try:
                    return float(v)
                except (ValueError, TypeError):
                    return None

            def txt(col):
                v = ws.cell(row=row_idx, column=col).value
                return str(v).strip() if v else None

            torre_id = tower_map.get(est)

            record = {
                "torre_id": torre_id,
                "linha_id": linha_id,
                "est_codigo": est,
                "vao_frente_m": num(2),
                "largura_m": num(3),
                "map_mec_extensao": num(4),
                "map_mec_largura": num(5),
                "map_man_extensao": num(6),
                "map_man_largura": num(7),
                "exec_mec_extensao": num(8),
                "exec_mec_largura": num(9),
                "exec_man_extensao": num(10),
                "exec_man_largura": num(11),
                "roco_concluido": txt(13) == "SIM" if txt(13) else False,
                "atende": txt(14),
                "area_manual": num(15),
                "area_mecanizado": num(16),
                "desconto_seletivo": num(17),
                "conferencia_vao": num(18),
                "area_manual_m2": num(19),
                "area_mecanizado_m2": num(20),
                "desconto_seletivo_m2": num(21),
                "numeracao_ggt": int(num(22)) if num(22) else None,
                "mapeamento_ggt": txt(23),
                "codigo_ggt_execucao": txt(24),
                "descricao_servico": txt(25),
                "prioridade": txt(26),
                "corte_arvores_25cm": int(num(27)) if num(27) else 0,
                "corte_arvores_15cm": num(28) or 0,
                "campanha_roco": campanha_roco,
                "nome_linha_planilha": lt_name or None,
            }

            # Handle date column (L, col 12)
            date_val = ws.cell(row=row_idx, column=12).value
            if date_val:
                try:
                    if hasattr(date_val, 'strftime'):
                        record["data_conclusao"] = date_val.strftime("%Y-%m-%d")
                    else:
                        record["data_conclusao"] = str(date_val)
                except Exception:
                    pass

            try:
                supabase.table("mapeamento_supressao").upsert(
                    record, on_conflict="linha_id,est_codigo,campanha_roco"
                ).execute()
                imported += 1
                results.append({"est": est, "torre_id": torre_id, "status": "ok"})
            except Exception as e:
                errors += 1
                results.append({"est": est, "status": "error", "error": str(e)})

        return {
            "success": True,
            "sheet_name": ws.title,
            "lt_name": lt_name,
            "linha_id": linha_id,
            "total_rows": imported + skipped + errors,
            "imported": imported,
            "skipped": skipped,
            "errors": errors,
            "towers_matched": sum(1 for r in results if r.get("torre_id")),
            "towers_unmatched": sum(1 for r in results if not r.get("torre_id")),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/batch-analyze")
async def batch_analyze(photo_ids: List[str]):
    """Trigger analysis for multiple photos. Returns job status."""
    if not analyzer or not supabase:
        raise HTTPException(status_code=503, detail="Service not ready")

    results = []
    for photo_id in photo_ids:
        try:
            # Get photo URL from DB
            foto = supabase.table("fotos").select("caminho_storage").eq("id", photo_id).single().execute()
            if foto.data:
                storage_path = foto.data["caminho_storage"]
                image_url = f"{SUPABASE_URL}/storage/v1/object/public/fotos-inspecao/{storage_path}"
                result = await analyzer.analyze_image(image_url, photo_id)

                # Save to DB
                supabase.table("ai_analysis").upsert({
                    "foto_id": photo_id,
                    "vegetation_detected": result.vegetation_detected,
                    "vegetation_score": result.vegetation_score,
                    "fire_signs": result.fire_signs,
                    "fire_score": result.fire_score,
                    "structural_issue": result.structural_issue,
                    "anomaly_type": result.anomaly_type,
                    "severity_score": result.severity_score,
                    "confidence": result.confidence,
                    "quality_blur": result.quality.blur_score,
                    "quality_exposure": result.quality.exposure_score,
                    "summary": result.summary,
                    "model_version": result.model_version,
                }, on_conflict="foto_id").execute()

                results.append({"photo_id": photo_id, "status": "ok", "severity": result.severity_score})
        except Exception as e:
            results.append({"photo_id": photo_id, "status": "error", "error": str(e)})

    return {"processed": len(results), "results": results}


@app.post("/parse-roco-text", response_model=ParseRocoResponse)
async def parse_roco_text(request: ParseRocoRequest):
    """Extrai metragens e tipos de roço a partir de um texto livre via AI."""
    if not analyzer:
        raise HTTPException(status_code=503, detail="Analyzer not initialized")
    
    try:
        result = await analyzer.parse_roco_text(request.texto, request.vao_m)
        return ParseRocoResponse(segmentos=[RocoSegmentModel(**seg) for seg in result.get('segmentos', [])])
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


class MoondreamAnnotateRequest(BaseModel):
    image_url: str
    photo_id: Optional[str] = None
    torre_codigo: Optional[str] = None
    vao_total_m: float = 100.0
    largura_m: float = 40.0  # Right-of-way strip width in meters
    segments: list  # List of {tipo, inicio, fim}


@app.post("/annotate-moondream")
async def annotate_moondream(request: MoondreamAnnotateRequest):
    """Use Moondream 3 to detect the vegetation corridor and overlay mapping segment distances."""
    if not analyzer:
        raise HTTPException(status_code=503, detail="Analyzer not initialized")

    try:
        # Fetch torre_codigo from DB if not provided
        torre_codigo = request.torre_codigo or ""
        if not torre_codigo and request.photo_id and supabase:
            try:
                foto = supabase.table("fotos").select(
                    "torres(codigo_torre)"
                ).eq("id", request.photo_id).maybe_single().execute()
                if foto.data and foto.data.get("torres"):
                    torre_codigo = foto.data["torres"].get("codigo_torre", "")
            except Exception as e:
                print(f"[annotate-moondream] DB query error: {e}")

        # Download image
        img = await analyzer.download_image(request.image_url)

        # Annotate with Moondream bounding boxes + segment overlay
        annotated_bytes = await analyzer.annotate_with_moondream(
            img,
            segments=request.segments,
            vao_total_m=request.vao_total_m,
            largura_m=request.largura_m,
            torre_codigo=torre_codigo,
        )

        from fastapi.responses import Response
        return Response(
            content=annotated_bytes,
            media_type="image/jpeg",
            headers={"X-Torre": torre_codigo},
        )
    except Exception as e:
        error_msg = f"[annotate-moondream] ERROR: {e}"
        print(error_msg)
        import traceback
        logging.error(f"{error_msg}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("AI_SERVICE_PORT", "8000"))
    uvicorn.run("main:app", host="0.0.0.0", port=port, reload=True)
